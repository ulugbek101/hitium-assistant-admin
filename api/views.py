from json import dumps
from urllib.parse import quote
from datetime import datetime, timedelta, date
import calendar
from collections import defaultdict

from django.utils import timezone
from django.db.models import Q, F, Prefetch
from django.http import JsonResponse
from django.contrib.auth import get_user_model
from django.http import HttpResponse

from rest_framework.views import csrf_exempt
from rest_framework.decorators import api_view
from rest_framework.response import Response

import openpyxl
from openpyxl.styles import Alignment, PatternFill

from django.db import transaction
from api.models import User, Specialization, Task, FinishedWork, BotUser, Attendance, Day
from api.serializers import TaskSerializer


User = get_user_model()



@transaction.atomic
def register_user_in_db(data, files):
    phone_number = data.get("phone_number")
    telegram_id = data.get("telegram_id")

    # 1. Resolve specialization
    specialization = None
    specialization_name = data.get("specialization")

    if specialization_name:
        specialization, _ = Specialization.objects.get_or_create(
            name=specialization_name
        )

    # 2. Find user by unique fields
    user = (
        User.objects
        .filter(phone_number=phone_number)
        .first()
        or
        User.objects.filter(telegram_id=telegram_id).first()
    )

    # 3. Create user if not exists
    if not user:
        user = User.objects.create(
            telegram_id=telegram_id,
            phone_number=phone_number,
            first_name=data.get("first_name"),
            last_name=data.get("last_name"),
            middle_name=data.get("middle_name"),
            born_year=datetime.strptime(data.get("born_year"), "%Y-%m-%d").date(),
            type_of_document=data.get("type_of_document"),
            card_number=data.get("card_number"),
            card_holder_name=data.get("card_holder_name"),
            tranzit_number=data.get("tranzit_number"),
            bank_name=data.get("bank_name"),
            specialization=specialization,
            passport_photo=files.get("passport_photo"),
            id_card_photo1=files.get("id_card_photo1"),
            id_card_photo2=files.get("id_card_photo2"),
        )
        return user, True

    # 4. Update existing user (only provided fields)
    user.telegram_id = telegram_id or user.telegram_id
    user.first_name = data.get("first_name") or user.first_name
    user.last_name = data.get("last_name") or user.last_name
    user.middle_name = data.get("middle_name") or user.middle_name
    user.type_of_document = data.get("type_of_document") or user.type_of_document
    user.card_number = data.get("card_number") or user.card_number
    user.card_holder_name = data.get("card_holder_name") or user.card_holder_name
    user.tranzit_number = data.get("tranzit_number") or user.tranzit_number
    user.bank_name = data.get("bank_name") or user.bank_name
    user.born_year = datetime.strptime(data.get("born_year"), "%Y-%m-%d").date() or user.born_year
    user.specialization = specialization or user.specialization

    # 5. Update files only if sent
    if files.get("passport_photo"):
        user.passport_photo = files["passport_photo"]

    if files.get("id_card_photo1"):
        user.id_card_photo1 = files["id_card_photo1"]

    if files.get("id_card_photo2"):
        user.id_card_photo2 = files["id_card_photo2"]

    user.save()

    return user, False


@csrf_exempt
def register_user(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    # Text data
    user_data = request.POST.dict()

    # Files
    user_files = request.FILES

    register_user_in_db(user_data, user_files)

    return JsonResponse({"status": "ok"})


@api_view(["GET"])
def get_tasks(request):
    telegram_id = request.GET.get("telegram_id")

    tasks = Task.objects.filter(
        Q(brigades__foreman__telegram_id=telegram_id) |  # foreman of brigade
        Q(brigades__workers__telegram_id=telegram_id)    # any worker in brigade
    ).distinct()

    tasks = TaskSerializer(tasks, many=True).data

    return Response(data=tasks)


@api_view(["GET"])
def get_specializations(request):
    specializations = (
        Specialization.objects
        .values_list("name", flat=True)
        .distinct()
    )
    return Response(list(specializations))


def tasks(request):
    result = Task.objects.count()

    if result >= 100:
        result = "99+"

    return result


def finished_works(request):
    result = FinishedWork.objects.filter(is_done=False).count()

    if result >= 100:
        result = "99+"

    return  result


def download_attendance_report(request):
    period = request.GET.get("period", "current_month")
    today = timezone.now()

    full_date = timezone.now()

    year = full_date.year
    month = full_date.month

    if period == "prev_month":
        month -= 1
        if month < 1:
            month = 12
            year -= 1

    # Days in the month
    _, last_day = calendar.monthrange(year, month)
    month_days = [date(year, month, day) for day in range(1, last_day + 1)]

    # Fetch all Day objects for the month
    day_objs = Day.objects.filter(date__year=year, date__month=month)
    
    # Prefetch Attendance for all users for the month
    attendances_qs = Attendance.objects.filter(day__in=day_objs)
    users = BotUser.objects.prefetch_related(
        Prefetch('attendance_set', queryset=attendances_qs, to_attr='month_attendance')
    ).order_by('first_name')

    # Prepare workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    title = f"{month if month > 9 else '0' + str(month)}.{year}"
    ws.title = title

    header = ["ФИО сотрудника"] + [f"{d.day if d.day > 9 else '0' + str(d.day)}" for d in month_days] + ["Всего за месяц", "Рабочие дни", "Пропущенные дни"]
    ws.append(header)
    
    for col in range(1, len(header) + 1):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    today_date = timezone.now().date()
    sunday_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # red bg for Sunday
    missed_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # orange bg for missed working day

    for user in users:
        row = [f"{user.first_name or ''} {user.last_name or ''} {user.middle_name or ''}"]
        total_seconds = 0
        worked_days = 0
        missed_days = 0
        attendance_map = {
            a.day.date: a
            for a in getattr(user, 'month_attendance', [])
        }

        for day_date in month_days:
            cell_value = "-"
            # att = next((a for a in getattr(user, 'month_attendance', []) if a.day.date == day_date), None)
            att = attendance_map.get(day_date)

            is_sunday = day_date.weekday() == 6
            is_past_or_today = day_date <= today_date  # includes today
            is_future = day_date > today_date

            if att and att.start_time:
                if not att.end_time:
                    cell_value = f"В процессе"
                    worked_seconds = 0
                else:  
                    worked_seconds = (datetime.combine(day_date, att.end_time) - datetime.combine(day_date, att.start_time)).seconds
                    hours = worked_seconds // 3600
                    minutes = (worked_seconds % 3600) // 60
                    cell_value = f"{hours} ч. {minutes} мин."

                total_seconds += worked_seconds
                worked_days += 1
            elif not is_sunday and is_past_or_today:
                # Count as missed working day only if day is past or today
                missed_days += 1

            row.append(cell_value)

        # Subtract the time workers spent for lunch
        worked_seconds -= 3600
        
        total_hours = total_seconds // 3600
        total_minutes = (total_seconds % 3600) // 60
        row.append(f"{total_hours} ч. {total_minutes} мин.")
        row.append(worked_days)
        row.append(missed_days)

        ws.append(row)

        # Center all cells and apply coloring
        for col_idx, day_date in enumerate(month_days, start=2):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.alignment = Alignment(horizontal="center")

            if day_date <= today_date:  # only color past days and today
                if day_date.weekday() == 6:
                    cell.fill = sunday_fill
                elif cell.value == "-":
                    cell.fill = missed_fill

        # Center total/worked/missed columns
        for col_idx in range(len(row) - 2, len(row) + 1):
            ws.cell(row=ws.max_row, column=col_idx).alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        col[0].column_letter
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    filename = f"Тебель рабочего времени за {month:02}.{year}.xlsx"
    quoted_filename = quote(filename)  # URL-encode UTF-8
    
    # Return Excel response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quoted_filename}"
    wb.save(response)
    return response
